import tkinter as tk
from tkinter import ttk, filedialog
from PIL import ImageTk, Image
from openpyxl import load_workbook
import os

TREEVIEW_COLUMN_WIDTH = 75
TREEVIEW_COLUMN_MIN_WIDTH = 50

class MPIBProblemExcelLoading(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent)

        self.grid_rowconfigure(2, weight=1)
        #self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.wb = None

        lblInfo = tk.Label(self, anchor="w", justify=tk.LEFT, fg='green')
        lblInfo.config(text="""1. 엑셀 파일에 있는 자료로부터 문제를 등록합니다. 그림 파일은 '*.jpg'와 '*.png'만 가능합니다.
2. 첫 행은 제목행으로 사용합니다. 'A1' 셀에는 문제가 들어 있는 폴더명이 들어 있어야 합니다. 그렇지 않으면 다른 내용을 아예 표시하지 않습니다.
3. 첫 열에는 2행부터 문제 파일의 이름이 있어야 합니다. 빈 행이 있을 때까지 읽어 들입니다. (답 파일과 문제 파일은 시트에서 별도로 나열하지 않습니다.)
4. 답 파일의 이름은 '(문제 파일 이름)_ans', 풀이 파일의 이름은 '(문제 파일 이름)_sol'로 되어 있어야 하며, '4'에 나오는 '주관식 답(그림)의 존재 여부'와 '풀이(그림)의 존재 여부'에 1의 값을 부여해야 합니다.

  예) 문제: problem.jpg, 답: problem_ans.png, 풀이: problem_sol.jpg (* 문제, 답, 풀이 파일들 사이에 jpg와 png를 섞어 쓰는 것은 가능)

4. 두 번째 열부터 출처, 문제 유형, 문제 단 수, 난이도, 답 유형, 객관식 답, 주관식 답(텍스트), 주관식 답(그림) 존재 여부, 풀이 존재 여부, 풀이 단 수의 순서로 자료가 들어 있어야 합니다.
5. '주관식 답(텍스트)'에도 값이 있고 '주관식 답(그림)의 존재 여부'도 1로 되어 있으면 '주관식 답(텍스트)에 있는 값만 적용됩니다.""")
        lblInfo.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

        frmHeader = tk.Frame(self)
        frmHeader.grid_columnconfigure(1, weight=1)
        frmHeader.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

        btnSelectExcelFile = tk.Button(frmHeader, text='엑셀 파일 선택', bg='yellow', command=self.selectExcelFile)
        btnSelectExcelFile.grid(row=0, column=0)
        
        self.lblSelectExcelFile = tk.Label(frmHeader)
        self.lblSelectExcelFile.grid(row=0, column=1, padx=5, sticky="w")

        frmSheets = tk.Frame(self) # 시트 목록을 보여준다.
        frmSheets.grid_rowconfigure(1, weight=1)
        #frmSheets.grid_columnconfigure(0, weight=1)
        frmSheets.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        self.lblSheet = tk.Label(frmSheets)
        self.lblSheet.grid(row=0, column=0, columnspan=2, sticky="ew")

        self.lstSheet = tk.Listbox(frmSheets, width=25, selectmode=tk.SINGLE)
        self.lstSheet.grid(row=1, column=0, sticky="ns")
        self.lstSheet.bind('<<ListboxSelect>>', self.showSheetInfo)

        yScroll = ttk.Scrollbar(frmSheets, orient=tk.VERTICAL) # 수직 스크롤바 만들기
        yScroll.configure(command=self.lstSheet.yview)
        self.lstSheet.configure(yscrollcommand=yScroll.set)
        yScroll.grid(row=1, column=1, sticky="ns")

        frmSheet = tk.Frame(self) # 시트 내용의 일부를 보여준다.
        frmSheet.grid_rowconfigure(0, weight=1)
        frmSheet.grid_columnconfigure(0, weight=1)
        frmSheet.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")

        self.problemSheet = ttk.Treeview(frmSheet, selectmode='browse')
        self.problemSheet.grid(row=0, column=0, sticky="nsew")

        self.problemSheet.column('#0', anchor='center', minwidth=200)
        self.problemSheet.config(columns=('sourceID', 'problemTypeID', 'columns', 'difficulty', 'ansType',
                                          'objAns', 'subjAns', 'subjAnsImageExists', 'solImageExists', 'solColumns',))
        self.problemSheet.column('sourceID', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('sourceID', text='출처') # 관리자 전용
        self.problemSheet.column('problemTypeID', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('problemTypeID', text='문제 유형')
        self.problemSheet.column('columns', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('columns', text='문제 단 수')
        self.problemSheet.column('difficulty', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('difficulty', text='난이도')
        self.problemSheet.column('ansType', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('ansType', text='답 유형')
        self.problemSheet.column('objAns', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('objAns', text='객관식 답')
        self.problemSheet.column('subjAns', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('subjAns', text='주관식 답(텍스트)')
        self.problemSheet.column('subjAnsImageExists', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('subjAnsImageExists', text='주관식 답(그림) 존재 여부')
        self.problemSheet.column('solImageExists', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('solImageExists', text='풀이 존재 여부')
        self.problemSheet.column('solColumns', anchor="center", minwidth=TREEVIEW_COLUMN_MIN_WIDTH, width=TREEVIEW_COLUMN_WIDTH)
        self.problemSheet.heading('solColumns', text='풀이 단 수')

        trvYScroll = ttk.Scrollbar(frmSheet, orient=tk.VERTICAL) # 수직 스크롤바 만들기
        trvYScroll.configure(command=self.problemSheet.yview)
        self.problemSheet.configure(yscrollcommand=trvYScroll.set)
        trvYScroll.grid(row=0, column=1, sticky="ns")

        trvXScroll = ttk.Scrollbar(frmSheet, orient=tk.HORIZONTAL) # 수평 스크롤바 만들기
        trvXScroll.configure(command=self.problemSheet.xview)
        self.problemSheet.configure(xscrollcommand=trvXScroll.set)
        trvXScroll.grid(row=1, column=0, sticky="ew")

        self.problemSheet.bind('<Double-1>', self.checkAProblem) # 더블 클릭 시 개별 확인

    def selectExcelFile(self):
        excelFilename = filedialog.askopenfilename(filetypes=(("Excel 파일", "*.xls; *.xlsx; *.xlsm"),))

        if excelFilename != '': # 선택한 파일이 있으면
            # 우선 기존 내용 삭제
            self.lstSheet.delete(0, 'end') # 시트 목록부터 삭제
            
            for i in self.problemSheet.get_children(): # 트리 내용도 삭제
                self.problemSheet.delete(i)
            self.problemSheet.heading('#0', text='')

            # 새 내용 기록
            self.lblSelectExcelFile.config(text=excelFilename)
            self.lblSheet.config(text='시트를 선택하세요.', fg='green')

            self.wb = load_workbook(excelFilename)
            
            for ws in self.wb.sheetnames:
                self.lstSheet.insert(tk.END, ws)
    
    def showSheetInfo(self, event):        
        for i in self.problemSheet.get_children(): # 우선 기존 내용 삭제
            self.problemSheet.delete(i)
            
        ws = self.wb[self.lstSheet.get(self.lstSheet.curselection())]
        
        if ws.cell(row=1, column=1).value != None: # 그림이 들어 있는 폴더 표시
            self.problemSheet.heading('#0', text=ws.cell(row=1, column=1).value)

            rowCount = 0
            firstColumn = ws['A']

            for aCell in firstColumn:
                if aCell.value is None:
                    break
                else:
                    rowCount += 1
            
            for row in ws.iter_rows(min_row=2, max_row=rowCount, max_col=11):
                value = []

                for i in range(0, 11):
                    value.append('' if row[i].value is None else row[i].value)

                #print(values)
                self.problemSheet.insert('', 'end', value[0], text=value[0], values=(value[1], value[2], value[3], value[4], value[5], value[6],
                                         value[7], value[8], value[9], value[10]))
        else:
            self.problemSheet.heading('#0', text='')

    def checkAProblem(self, event):
        region = self.problemSheet.identify('region', event.x, event.y)

        if region == 'tree' or region == 'cell':
            aFolder = self.problemSheet.heading('#0')['text']

            if not os.path.exists(aFolder)):
                break

            aFile = self.problemSheet.identify('item', event.x, event.y)
            aPath = os.path.join(aFolder, aFile)

            #aWin = tk.Toplevel()
            #aWin.title(aPath)
            #aWin.grab_set()
            #aWin.grid_rowconfigure(0, weight=1)
            #aWin.grid_columnconfigure(0, weight=1)

            #aLabel = tk.Label(aWin, width=400, height=300)
            #self.img = Image.open(aPath)
            #self.original = ImageTk.PhotoImage(self.img)
            #aLabel.config(image=self.original)
            #aLabel.grid(row=0, column=0)

        

if __name__ == '__main__':
    root = tk.Tk()
    root.title("엑셀로부터 문제 등록하기")
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    test = MPIBProblemExcelLoading(root)
    test.grid(sticky="nsew")

    root.mainloop()